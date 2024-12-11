import os
import random
import locale
import pytz
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes
from openpyxl import Workbook, load_workbook

class NequiBot:
    def __init__(self):
        self.template_path = "template_nequi.png"
        self.font_regular = "Manrope-Regular.ttf"
        self.font_sizes = {
            'nombre': 29,
            'numero': 29,
            'fecha': 29,
            'cantidad': 29,
            'cantidad_secondary': 27.5,
            'random_code': 29
        }
        self.user_file = "usuarios_registrados.txt"
        self.allowed_users = []
        self.pending_requests = {}
        self.admin_user_id = 6374048796  # ID del administrador principal
        self.command_history = []
        self.excel_file = "comandos_nequi.xlsx"
        self.generated_images = []

        self.ensure_user_file_exists()
        self.load_users()
        self.init_excel()

        # Asegurarse de que el admin esté siempre en la lista de usuarios permitidos
        if self.admin_user_id not in self.allowed_users:
            self.allowed_users.append(self.admin_user_id)
            self.save_users()

    def get_font(self, size):
        """Obtiene una fuente con el tamaño especificado."""
        return ImageFont.truetype(self.font_regular, size)

    def generate_random_code(self):
        """Genera un código aleatorio."""
        random_code = "M" + ''.join([str(random.randint(0, 9)) for _ in range(7)])
        return random_code

    def save_to_excel(self, user_id, nombre, numero, cantidad, fecha, random_code):
        """Guarda los datos en el archivo Excel."""
        wb = load_workbook(self.excel_file)
        ws = wb.active
        ws.append([user_id, nombre, numero, cantidad, fecha, random_code])  
        wb.save(self.excel_file)

    def init_excel(self):
        """Inicializa el archivo Excel si no existe y crea las cabeceras."""
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Usuario ID", "Nombre", "Número", "Cantidad", "Fecha", "Código aleatorio"])  # Cabeceras
            wb.save(self.excel_file)

    def ensure_user_file_exists(self):
        """Crea el archivo de usuarios registrados si no existe."""
        if not os.path.exists(self.user_file):
            with open(self.user_file, 'w') as f:
                pass  # Crea el archivo vacío.

    def load_users(self):
        """Carga los usuarios registrados desde el archivo."""
        with open(self.user_file, 'r') as f:
            self.allowed_users = [int(line.strip()) for line in f if line.strip().isdigit()]

    def save_users(self):
        """Guarda los usuarios registrados en el archivo."""
        with open(self.user_file, 'w') as f:
            f.writelines(f"{user_id}\n" for user_id in self.allowed_users)

    async def start_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if user_id in self.allowed_users:
            await update.message.reply_text(
                "Bot creado por @munano13.\n"
                "Si quieres acceso, escribe a @munano13 \n"
                "y por favor poner tu @ público, de lo contrario serás rechazado automaticamente" 
            )
        else:
            await update.message.reply_text(
                "Bot creado por @munano13.\n"
                "Para registrarse, favor colocar /registrarse (ANTES DE REGISTRATE HABLA CON @munano13 \n"
                "y por favor poner tu @ público, de lo contrario serás rechazado automaticamente)"
            )

    async def registrarse_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        username = update.effective_user.username

        if user_id in self.allowed_users:
            await update.message.reply_text("Ya estás registrado en el sistema.")
            return

        if user_id not in self.pending_requests:
            self.pending_requests[user_id] = {
                'username': username,
                'user_id': user_id
            }

            await update.message.reply_text("Solicitud de aprobación enviada. Espera la confirmación del administrador. Si tienes el @ en privado serás rechazado automaticamente")
            try:
                admin_chat = await context.bot.get_chat(self.admin_user_id)
                await admin_chat.send_message(f"Nuevo usuario solicitando aprobación:\nID de Usuario: {user_id}\nNombre de usuario: @{username}")
            except Exception as e:
                await update.message.reply_text(f"Error al notificar al administrador: {e}")
        else:
            await update.message.reply_text("Ya has enviado una solicitud de aprobación.")

    async def aceptar_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id != self.admin_user_id:
            await update.message.reply_text("No tienes permiso para aceptar usuarios.")
            return

        if not context.args:
            await update.message.reply_text("Uso: /aceptar <ID_DEL_USUARIO>")
            return

        user_id = int(context.args[0])

        if user_id in self.allowed_users:
            await update.message.reply_text("El usuario ya está registrado.")
            return

        if user_id in self.pending_requests:
            self.allowed_users.append(user_id)
            self.save_users()
            del self.pending_requests[user_id]
            await update.message.reply_text("Usuario aceptado correctamente.")
            try:
                await context.bot.send_message(chat_id=user_id, text="Has sido aceptado. Ya puedes usar el bot.")
            except Exception as e:
                await update.message.reply_text(f"Error notificando al usuario: {e}")
        else:
            await update.message.reply_text("No hay solicitud pendiente para este usuario.")

    async def rechazar_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id != self.admin_user_id:
            await update.message.reply_text("No tienes permiso para rechazar usuarios.")
            return

        if not context.args:
            await update.message.reply_text("Uso: /rechazar <ID_DEL_USUARIO>")
            return

        user_id = int(context.args[0])

        if user_id in self.pending_requests:
            del self.pending_requests[user_id]
            await update.message.reply_text(f"Solicitud de usuario con ID {user_id} rechazada.")
            try:
                await context.bot.send_message(chat_id=user_id, text="Tu solicitud de registro ha sido rechazada.")
            except Exception as e:
                await update.message.reply_text(f"Error notificando al usuario: {e}")
        else:
            await update.message.reply_text(f"No hay solicitud pendiente para el usuario con ID {user_id}.")

    async def consultas_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id != self.admin_user_id:
            await update.message.reply_text("No tienes permiso para ver las consultas.")
            return

        if not self.allowed_users:
            await update.message.reply_text("No hay usuarios registrados.")
            return

        users_list = "\n".join(map(str, self.allowed_users))
        await update.message.reply_text(f"Usuarios registrados:\n{users_list}")

    async def borrar_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Elimina a un usuario de la lista de usuarios registrados."""
        if update.effective_user.id != self.admin_user_id:
            await update.message.reply_text("No tienes permiso para borrar usuarios.")
            return

        if not context.args:
            await update.message.reply_text("Uso: /borrar <ID_DEL_USUARIO>")
            return

        try:
            user_id = int(context.args[0])
        except ValueError:
            await update.message.reply_text("Por favor, proporciona un ID válido.")
            return

        if user_id in self.allowed_users:
            self.allowed_users.remove(user_id)
            self.save_users()
            await update.message.reply_text(f"Usuario con ID {user_id} eliminado correctamente.")
            try:
                await context.bot.send_message(chat_id=user_id, text="Has sido eliminado de la lista de usuarios registrados.")
            except Exception as e:
                await update.message.reply_text(f"No se pudo notificar al usuario: {e}")
        else:
            await update.message.reply_text("El usuario no está en la lista de registrados.")

    async def registro_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        # Compatibilidad con el comando de registro anterior
        if update.effective_user.id != self.admin_user_id:
            await update.message.reply_text("No tienes permiso para ver el historial de comandos.")
            return

        if not self.command_history:
            await update.message.reply_text("No hay registros de comandos aún.")
            return

        for entry in self.command_history:
            user_id = entry['user_id']
            nombre = entry['nombre']
            numero = entry['numero']
            cantidad = entry['cantidad']
            fecha = entry['fecha']
            await update.message.reply_text(f"Usuario ID: {user_id}\nNombre: {nombre}\nNúmero: {numero}\nCantidad: {cantidad}\nFecha: {fecha}")

    async def nequi_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id

        # Bypass para el administrador principal
        if user_id == self.admin_user_id:
            pass
        # Verificar si el usuario está registrado
        elif user_id not in self.allowed_users:
            await update.message.reply_text("No tienes permiso para usar este bot. Escribe /registrarse para solicitar acceso.")
            return

        try:
            if len(context.args) < 3:
                await update.message.reply_text("Uso: /nequi <NOMBRE COMPLETO> <NÚMERO> <CANTIDAD>")
                return

            # Recolectar los argumentos
            nombre = " ".join(context.args[:-2])  
            numero = context.args[-2]  
            cantidad = context.args[-1]  

            if not numero.isdigit():
                await update.message.reply_text("El número debe ser válido y contener solo dígitos.")
                return

            cantidad_sin_puntos = cantidad.replace('.', '')  
            if not cantidad_sin_puntos.isdigit():
                await update.message.reply_text("La cantidad debe ser un número válido.")
                return

            cantidad_formateada = f"$ {int(cantidad_sin_puntos):,}".replace(',', '.') + ",00"

            locale.setlocale(locale.LC_TIME, 'es_CO.utf8')
            tz_col = pytz.timezone('America/Bogota')
            fecha_actual = datetime.now(tz_col).strftime("%d de %B de %Y,  %I:%M %p")
            fecha_actual = fecha_actual.replace("AM", "a. m.").replace("PM", "p. m.")
            fecha_actual = fecha_actual.encode('latin1').decode('utf-8')

            img = Image.open(self.template_path)
            draw = ImageDraw.Draw(img)

            nombre_font = self.get_font(self.font_sizes['nombre'])
            numero_font = self.get_font(self.font_sizes['numero'])
            fecha_font = self.get_font(self.font_sizes['fecha'])
            cantidad_font = self.get_font(self.font_sizes['cantidad'])
            cantidad_secondary_font = ImageFont.truetype("Manrope-SemiBold.ttf", int(self.font_sizes['cantidad_secondary']))

            coords = {
                'nombre': (61.5, 325),
                'numero': (60.9, 553),
                'fecha': (63, 660),
                'cantidad': (60.9, 773),
                'cantidad_secondary': (145, 965),
                'random_code': (60, 440)
            }

            draw.text(coords['nombre'], nombre.upper(), font=nombre_font, fill='#180b1b')
            draw.text(coords['numero'], numero, font=numero_font, fill='#180b1b')
            draw.text(coords['fecha'], fecha_actual, font=fecha_font, fill='#180b1b')
            draw.text(coords['cantidad'], cantidad_formateada, font=cantidad_font, fill='#180b1b')
            draw.text(coords['cantidad_secondary'], cantidad_formateada, font=cantidad_secondary_font, fill='#808080')

            random_code = self.generate_random_code()
            random_code_font = self.get_font(self.font_sizes['random_code'])

            draw.text(coords['random_code'], random_code, font=random_code_font, fill='#180b1b')

            temp_path = f"temp_{update.effective_chat.id}.png"
            img.save(temp_path, format='PNG', quality=95)

            self.generated_images.append({
                'user_id': user_id,
                'image_path': temp_path
            })

            self.command_history.append({
                'user_id': user_id,
                'nombre': nombre,
                'numero': numero,
                'cantidad': cantidad_formateada,
                'fecha': fecha_actual,
                'random_code': random_code
            })

            self.save_to_excel(user_id, nombre, numero, cantidad_formateada, fecha_actual, random_code)

            await update.message.reply_document(document=open(temp_path, 'rb'))
            os.remove(temp_path)

        except Exception as e:
            await update.message.reply_text(f"Error: {str(e)}")

def main():
    """Arranca el bot y agrega los comandos."""
    bot = NequiBot()
    application = Application.builder().token("7403954410:AAGFt7KlZWoxJ65n3PZWlioYJjsZt4iIVRU").build()

    # Agregar manejadores de comandos
    application.add_handler(CommandHandler("start", bot.start_command))
    application.add_handler(CommandHandler("registrarse", bot.registrarse_command))
    application.add_handler(CommandHandler("aceptar", bot.aceptar_command))
    application.add_handler(CommandHandler("rechazar", bot.rechazar_command))
    application.add_handler(CommandHandler("consultas", bot.consultas_command))
    application.add_handler(CommandHandler("borrar", bot.borrar_command))
    application.add_handler(CommandHandler("registro", bot.registro_command))
    application.add_handler(CommandHandler("nequi", bot.nequi_command))

    # Inicia el bot
    application.run_polling()

if __name__ == "__main__":
    main()
